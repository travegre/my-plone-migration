"""Definition of the seznam_zaposlenih content type (Dexterity)
"""

from zope.interface import implementer
from plone.dexterity.content import Container
from AccessControl import ClassSecurityInfo
from Products.CMFCore.permissions import ModifyPortalContent
from zope.lifecycleevent.interfaces import IObjectCreatedEvent, IObjectModifiedEvent

import plone.api
from dezurstva.produkti.interfaces import Iseznam_zaposlenih

import os
from zipfile import ZipFile
import mimetypes
from io import BytesIO
from DateTime.DateTime import DateTime

import openpyxl
import re
import csv
import logging
from Products.CMFPlone.utils import safe_unicode

log = logging.getLogger("Plone")


@implementer(Iseznam_zaposlenih)
class seznam_zaposlenih(Container):
    """Dexterity-based seznam_zaposlenih content type."""

    meta_type = "seznam_zaposlenih"
    portal_type = "seznam_zaposlenih"

    security = ClassSecurityInfo()

    security.declarePrivate('unpackArchive')
    def unpackArchive(self):
                
        
        field1_name = 'datoteka1'
        
        file1 = getattr(self, 'datoteka1', None)

        field2_name = 'datoteka2'
        
        file2 = getattr(self, 'datoteka2', None)

        def msg(message):
            plone.api.portal.show_message(message=message, request=self.REQUEST)
        
        if file1 and getattr(file1, 'data', None):       
            

            import datetime
            import re
            filename1 = getattr(file1, 'filename', 'file1.xlsx')
            
            # get the full binary from file
            try:
                chunk1 = file1.data if hasattr(file1, 'next') else None
                filedata1 = b''
                while chunk1 is not None:
                    filedata1 += chunk1.data if isinstance(chunk1.data, bytes) else chunk1.data.encode()
                    chunk1 = chunk1.next
            except AttributeError:
                filedata1 = file1.data if isinstance(file1.data, bytes) else file1.data.encode()

            wb = openpyxl.load_workbook(BytesIO(filedata1 if isinstance(filedata1, bytes) else filedata1.encode()))
            sh = wb.active

                       
            
            
            
            for i in range(sh.max_row):
                row = [sh.cell(row=i+1, column=j+1).value or "" for j in range(sh.max_column)]
                
                try: 
                    sifra = row[0]
                except:
                    sifra = ""
                try: 
                    delavec = row[1]
                except:
                    delavec = ""
                try: 
                    kontakt = row[2]
                except:
                    kontakt = ""
                
                preiskava = self.invokeFactory(
                                                                    
                                        type_name='Folder',
                                        id=sifra,
                                        title=delavec,
                                        description=kontakt                             
                                    )
                
        setattr(self, 'datoteka1', None)

        if file2 and getattr(file2, 'data', None): 

            import datetime
            import re
            filename2 = getattr(file2, 'filename', 'file2.xlsx')
            
            # get the full binary from file
            try:
                chunk2 = file2.data if hasattr(file2, 'next') else None
                filedata2 = b''
                while chunk2 is not None:
                    filedata2 += chunk2.data if isinstance(chunk2.data, bytes) else chunk2.data.encode()
                    chunk2 = chunk2.next               
            except AttributeError:
                filedata2 = file2.data if isinstance(file2.data, bytes) else file2.data.encode()

            wb = openpyxl.load_workbook(BytesIO(filedata2 if isinstance(filedata2, bytes) else filedata2.encode()))
            sh = wb.active
            

            
            for i in range(sh.max_row):
                row = [sh.cell(row=i+1, column=j+1).value or "" for j in range(sh.max_column)] 

                try: 
                    sifra = row[0]
                except:
                    sifra = ""
                try: 
                    delavec = row[1]
                except:
                    delavec = ""
                try: 
                    kontakt = row[2]
                except:
                    kontakt = ""
                try: 
                    email = row[3]
                except:
                    email = ""

                zadetek = self.portal_catalog(id=sifra)

                if zadetek:
                    zadetek = zadetek[0].getObject()
                    if email == '': email = '-'
                    zadetek.description = (kontakt + '|' + email)
                    pass  # debug print removed
                else:
                    if sifra != '':
                        if email == '': email = '-'
                        preiskava = self.invokeFactory(
                                                                        
                                            type_name='Folder',
                                            id=sifra,
                                            title=delavec,
                                            description=(kontakt + '|' + email)                        
                                        )
                    


        setattr(self, 'datoteka2', None)  


def on_object_created(obj, event):
    obj.unpackArchive()


def on_object_modified(obj, event):
    obj.unpackArchive()

