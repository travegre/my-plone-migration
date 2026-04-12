# coding=utf-8
"""Definition of the dezurstvo content type (Dexterity)
"""

from zope.interface import implementer
from plone.dexterity.content import Container
from AccessControl import ClassSecurityInfo
from Products.CMFCore.exceptions import BadRequest
from Products.CMFCore.permissions import ModifyPortalContent
from Products.CMFCore.utils import getToolByName
from Products.CMFCore.WorkflowCore import WorkflowException

from dezurstva.produkti.interfaces import Idezurstvo

import os
from zipfile import ZipFile
import mimetypes
from io import BytesIO
import base64
import openpyxl
import sys
import random
import DateTime
from datetime import datetime, timedelta


@implementer(Idezurstvo)
class dezurstvo(Container):
    """Dexterity-based dezurstvo content type."""

    meta_type = "dezurstvo"
    portal_type = "dezurstvo"

    security = ClassSecurityInfo()

    def getSampleVocabulary40(self):
        obj = self.portal.restrictedTraverse('dezurstva/seznam_zaposlenih')
        obj = obj.getFolderContents(contentFilter={'sort_on':'getObjPositionInParent'})             
                                 
        xx = []
        for rows in obj:
                if rows.id != "" and not rows.getObject().getExcludeFromNav():
                        xx.append((rows.id, rows.getObject().Title()))
        
        return xx

    security = ClassSecurityInfo()

    # this is used in the spremeni dezurstvo form
    def getSampleVocabulary50(self):
        obj = self.portal.restrictedTraverse('dezurstva/seznam_zaposlenih')
        obj = obj.getFolderContents(contentFilter={'sort_on':'getObjPositionInParent'})             
                                 
        xx = [('','')]
        for x in obj:
            if '|' in str(x.Description) and not x.getObject().getExcludeFromNav():
                xx.append((str(x.Description).split('|')[1] == '-' and x.id or str(x.Description).split('|')[1], x.Title))
        
        return xx

    security = ClassSecurityInfo()

    # -*- Your ATSchema to Python Property Bridges Here ... -*-

    def excel_export(self, first_day, last_day, oseba):
    
        dezurstva = self.portal_catalog(portal_type = "dezurstvo")
        #print dezurstva, first_day, last_day
        #for x in dezurstva:
        #    print x.id
        dezurstva = [x for x in dezurstva if 'undefined' not in x.id and x.id != 'objekt-za-seznam-zaposlenih-v-formi-spremeni-dezurstvo-ne-brisi' and DateTime.DateTime(x.id) >= first_day and DateTime.DateTime(x.id) <= last_day]

        if oseba:
            dezurstva = [self.portal_catalog(dezurni_zdravnik = oseba, id = x.id)[0] for x in dezurstva if len(self.portal_catalog(dezurni_zdravnik = oseba, id = x.id)) > 0]
        
        
        # create zip file
        #zipdat = BytesIO()
        #zf = ZipFile(zipdat, mode='w')

        # create excel
        workBookDocument = openpyxl.Workbook()
        docSheet1 = workBookDocument.active
        docSheet1.title = 'izpis'
        
        od = first_day.Date()
        od = '.'.join([od.split('/')[2], od.split('/')[1], od.split('/')[0]])
        do = last_day.Date()
        do = '.'.join([do.split('/')[2], do.split('/')[1], do.split('/')[0]])
        
        if oseba:
            person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + oseba).Title()
            docSheet1.cell(row=0+1, column=0+1, value='Izpis: ' + od + ' - ' + do + ' za osebo ' + person)
        else:
            docSheet1.cell(row=0+1, column=0+1, value='Izpis: ' + od + ' - ' + do)

        count = 3
        for j, i in enumerate(dezurstva):
            datum = '.'.join([i.id.split('-')[2], i.id.split('-')[1], i.id.split('-')[0]])
            i = i.getObject()
            #print(i.id)      
            lista = i.dezurni_zdravnik
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"dežurni zdravnik")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.nadzorni_zdravnik
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"nadzorni zdravnik")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.specializant
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"specializant na usposabljanju")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))
            '''
            lista = i.dezurni_tehnik
            print lista
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"dežurni tehnik")
                    person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                    docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    
            lista = i.izobrazevanje
            
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"izobraževanje")
                    person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                    docSheet1.cell(row=count+1, column=2+1, value=person.Title())
            '''
            lista = i.bakterioloski_tehnik
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"bakteriološki tehnik")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.inoqula_tehnik
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"inoqula tehnik")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))


            lista = i.sprejem_predpriprava_tehnik
            # To tukaj na silo dodajam ker sem preko plone admina odstranil, ampak naj se ve da so to šifro odpustil...
            if i.id == '2019-06-13':
                lista = ('418206', lista[0])
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"sprejem/predpriprava tehnik")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.vpis
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"vpis")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))


            lista = i.intervencijska_skupina
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"intervencijska_skupina")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))


            lista = i.BOR
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"BOR")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.HIV
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"HIV")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.HUM
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"HUM")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.PRZ
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"PRZ")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.IT
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"IT")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.KLM
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"KLM")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.KOV
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"KOV")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.VINVZV
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"VIN (VZV, EBV)")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.VINDIF
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"VIN (DIF ali PCR)")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.WHO
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"WHO")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))

            lista = i.kiestra
            #print("HEJ: ", lista)
            for h in lista:
                if not oseba or (oseba and oseba == h): 
                    count += 1
                    docSheet1.cell(row=count+1, column=0+1, value=datum)
                    docSheet1.cell(row=count+1, column=1+1, value=u"Kiestra")
                    try:
                        person = self.restrictedTraverse('dezurstva/seznam_zaposlenih/' + h)
                        docSheet1.cell(row=count+1, column=2+1, value=person.Title())
                    except:
                        docSheet1.cell(row=count+1, column=2+1, value=('sodelavec s šifro '+h+' ne obstaja več'))


	    #count += 1
            #docSheet1.cell(row=count+1, column=0+1, value='')


        dat = BytesIO()
        workBookDocument.save(dat)

        # output zip file
        r = self.portal.REQUEST  
        r.RESPONSE.setHeader("Content-type","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")        
        od = first_day.Date()
        od = '_'.join([od.split('/')[2], od.split('/')[1], od.split('/')[0]])
        do = last_day.Date()
        do = '_'.join([do.split('/')[2], do.split('/')[1], do.split('/')[0]])
        filename = 'izpis_' + od + '-' + do + '.xlsx'
        r.RESPONSE.setHeader("Content-disposition","attachment;filename=" + filename)    
        r.RESPONSE.write(dat.getvalue())
        
        return r.RESPONSE



    def ustvari_cel_teden(self):
        if self.klukca == 1:
            parent = self.aq_parent
            datum = self.title
            dezurstva = self.portal_catalog(portal_type="dezurstvo")
            dezurstva = [x for x in dezurstva if x.id != 'dezurstva/objekt-za-seznam-zaposlenih-v-formi-spremeni-dezurstvo-ne-brisi']
            dt = datetime.strptime(datum, '%Y-%m-%d')
            start = dt - timedelta(days=dt.weekday())
            for i in range(0, 7):
                try:
                    parent.invokeFactory(
                        type_name='dezurstvo',
                        id=start.strftime("%Y-%m-%d"),
                        title=start.strftime("%Y-%m-%d"),
                        BOR=self.BOR,
                        HIV=self.HIV,
                        HUM=self.HUM,
                        PRZ=self.PRZ,
                        IT=self.IT,
                        KLM=self.KLM,
                        KOV=self.KOV,
                        VINVZV=self.VINVZV,
                        VINDIF=self.VINDIF,
                        WHO=self.WHO,
                        kiestra=self.kiestra,
                    )
                except BadRequest:
                    for x in dezurstva:
                        dez = x.getObject()
                        if dez.id == start.strftime("%Y-%m-%d"):
                            dez.BOR = self.BOR
                            dez.HIV = self.HIV
                            dez.HUM = self.HUM
                            dez.PRZ = self.PRZ
                            dez.IT = self.IT
                            dez.KLM = self.KLM
                            dez.KOV = self.KOV
                            dez.VINVZV = self.VINVZV
                            dez.VINDIF = self.VINDIF
                            dez.WHO = self.WHO
                            dez.kiestra = self.kiestra
                try:
                    dez_obj = self.portal_catalog(portal_type="dezurstvo", id=start.strftime("%Y-%m-%d"))[0].getObject()
                    dez_obj.portal_workflow.doActionFor(dez_obj, "publish", comment="published programmatically")
                except WorkflowException:
                    pass
                start = start + timedelta(days=1)
            self.klukca = 0

    security.declareProtected(ModifyPortalContent, 'at_post_edit_script')
    def at_post_edit_script(self):
        self.ustvari_cel_teden()

    security.declareProtected(ModifyPortalContent, 'at_post_create_script')
    def at_post_create_script(self):
        self.ustvari_cel_teden()
