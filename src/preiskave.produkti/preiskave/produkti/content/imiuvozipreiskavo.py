# -*- coding: utf-8 -*-
"""Definition of the Imi uvozi preiskavo content type (Dexterity)
"""

from zope.interface import implementer
from plone.dexterity.content import Item
from AccessControl import ClassSecurityInfo
from Products.CMFCore.exceptions import BadRequest
from Products.CMFCore.permissions import ModifyPortalContent

from preiskave.produkti.interfaces import Iimiuvozipreiskavo

import logging
from io import BytesIO
import openpyxl

log = logging.getLogger("Plone")


@implementer(Iimiuvozipreiskavo)
class imiuvozipreiskavo(Item):
    """Dexterity-based imiuvozipreiskavo content type."""

    meta_type = "imiuvozipreiskavo"
    portal_type = "imiuvozipreiskavo"

    datoteka1 = None
    datoteka2 = None

    security = ClassSecurityInfo()
    security.declarePrivate('unpackArchive')

    def unpackArchive(self):
        file1 = getattr(self, 'datoteka1', None)
        file2 = getattr(self, 'datoteka2', None)
        filename1 = getattr(file1, 'filename', 'file1.xlsx')
        filename2 = getattr(file2, 'filename', 'file2.xlsx')

        msg = self.plone_utils.addPortalMessage

        if file1 and getattr(file1, 'data', None):
            msg('UVOZ PREISKAV:')
            stevec1 = 0
            stevec2 = 0
            stevec3 = 0

            # Open first file (preiskave data)
            dat1 = BytesIO(file1.data if isinstance(file1.data, bytes) else file1.data.encode())
            xl_workbook1 = openpyxl.load_workbook(dat1)
            xl_sheet1 = xl_workbook1.active

            # Open second file (vzorci data)
            dat2 = BytesIO(file2.data if isinstance(file2.data, bytes) else file2.data.encode())
            xl_workbook2 = openpyxl.load_workbook(dat2)
            xl_sheet2 = xl_workbook2.active

            msg('Stevilo vseh zaznanih vrstic v datoteki ' + filename1 + ': ' + str(xl_sheet1.max_row))
            msg('Stevilo vseh zaznanih vrstic v datoteki ' + filename2 + ': ' + str(xl_sheet2.max_row))

            vzorci = {}
            preiskave = {}
            preiskave2 = {}
            povezava = {}
            povezava_sin = {}
            povezava_nivo1 = {}
            povezava_nivo2 = {}
            povezava_nivo3 = {}
            povezava_nivo4 = {}
            povezava_nivo5 = {}
            povezava_lab = {}
            povezava_tel = {}
            povezava_odvzem = {}
            povezava_kolicina = {}
            povezava_transport = {}
            povezava_opomba = {}

            # SHRANIMO PODATKE OD PREISKAVE
            for row_idx in range(1, xl_sheet1.max_row):
                row = []
                for col_idx in range(0, xl_sheet1.max_column):
                    cell_obj = xl_sheet1.cell(row=row_idx + 1, column=col_idx + 1)
                    row.append(cell_obj.value)

                sifra = str(row[0]) if row[0] is not None else ''
                naziv = row[1]
                podrocje = row[2]
                sklop = row[3]
                sinonim = row[4]
                laboratoriji = row[5]
                metode = row[6]
                opis = row[7]
                opomba = row[8]
                trajanje = row[9]
                urnik = row[10]
                link_sop = row[11]
                nova_pre = str(row[12]) if row[12] is not None else ''
                nujna_pre = row[13]
                VzorecNaziv = row[14]
                VzorecSifra = str(row[15]) if row[15] is not None else ''
                PreiskavaVzorecNujno = row[16]
                Lab_1 = row[17] or ''
                Lab1 = row[18] or ''
                TelLab1 = str(row[19]) if row[19] is not None else ''
                Lab_2 = row[20] or ''
                Lab2 = row[21] or ''
                TelLab2 = str(row[22]) if row[22] is not None else ''
                Lab_3 = row[23] or ''
                Lab3 = ''
                TelLab3 = str(row[24]) if row[24] is not None else ''
                NavodiloZaOdvzem = row[25]
                NavodiloZaOdvzemVideoSifra = row[26]
                KolicinaInEmbalaza = row[27]
                EmbalazaSlikaSifra = row[28]
                NavodilaZaTransport = row[29]
                OpombaZaKlinika = row[30] if len(row) > 30 else ''

                try:
                    preiskave[sifra][0].append(podrocje)
                    preiskave[sifra][10].append(VzorecSifra)
                    preiskave[sifra][11].append(Lab_1 + '###' + Lab_2 + '###' + Lab_3)
                    preiskave[sifra][12].append(Lab1 + '###' + Lab2 + '###' + Lab3)
                    preiskave[sifra][13].append(TelLab1 + '###' + TelLab2 + '###' + TelLab3)
                    preiskave[sifra][14].append(NavodiloZaOdvzem)
                    preiskave[sifra][15].append(NavodiloZaOdvzemVideoSifra)
                    preiskave[sifra][16].append(KolicinaInEmbalaza)
                    preiskave[sifra][17].append(EmbalazaSlikaSifra)
                    preiskave[sifra][18].append(NavodilaZaTransport)
                    preiskave[sifra][19].append(OpombaZaKlinika)
                    preiskave[sifra][24].append(PreiskavaVzorecNujno)
                except KeyError:
                    preiskave[sifra] = [
                        [podrocje], sifra, naziv, metode, opis, sinonim, trajanje, urnik,
                        opomba, sklop, [VzorecSifra],
                        [Lab_1 + '###' + Lab_2 + '###' + Lab_3],
                        [Lab1 + '###' + Lab2 + '###' + Lab3],
                        [TelLab1 + '###' + TelLab2 + '###' + TelLab3],
                        [NavodiloZaOdvzem], [NavodiloZaOdvzemVideoSifra],
                        [KolicinaInEmbalaza], [EmbalazaSlikaSifra],
                        [NavodilaZaTransport], [OpombaZaKlinika],
                        laboratoriji, link_sop, nova_pre, nujna_pre, [PreiskavaVzorecNujno]
                    ]

                vzorci[VzorecSifra] = VzorecNaziv
                povezava[sifra] = []
                povezava_sin[sifra] = []
                povezava_nivo1[sifra] = []
                povezava_nivo2[sifra] = []
                povezava_nivo3[sifra] = []
                povezava_nivo4[sifra] = []
                povezava_nivo5[sifra] = []
                povezava_lab[sifra] = []
                povezava_tel[sifra] = []
                povezava_odvzem[sifra] = []
                povezava_kolicina[sifra] = []
                povezava_transport[sifra] = []
                povezava_opomba[sifra] = []

            # SHRANIMO PODATKE OD VZORCEV
            for row_idx in range(1, xl_sheet2.max_row):
                row = []
                for col_idx in range(0, xl_sheet2.max_column):
                    cell_obj = xl_sheet2.cell(row=row_idx + 1, column=col_idx + 1)
                    row.append(cell_obj.value)

                sifra_vzorca = str(row[0]) if row[0] is not None else ''
                nivo1 = row[1] or ''
                nivo2 = row[2] or ''
                nivo3 = row[3] or '' if len(row) > 3 else ''
                nivo4 = row[4] or '' if len(row) > 4 else ''
                nivo5 = row[5] or '' if len(row) > 5 else ''
                nivo6 = row[6] or '' if len(row) > 6 else ''

                if sifra_vzorca in vzorci:
                    vzorci[sifra_vzorca] = (
                        str(vzorci[sifra_vzorca]) + '######' +
                        nivo1 + '###' + nivo2 + '###' + nivo3 +
                        '###' + nivo4 + '###' + nivo5 + '###' + nivo6
                    )

            # ZACNEMO POSODABLJATI/USTVARJATI PREISKAVE
            preiskave210 = list(preiskave.keys())

            for row2 in preiskave210:
                row = preiskave[row2]
                novid = "preiskava_" + row2
                zadetek = self.portal_catalog(portal_type="imipreiskava", id=novid)

                preisk_vzorci = []
                preisk_vzorci_nivo1 = []
                preisk_vzorci_nivo2 = []
                preisk_vzorci_nivo3 = []
                preisk_vzorci_nivo4 = []
                preisk_vzorci_nivo5 = []
                preisk_vzorci_nivo6 = []
                row_vzorci = []

                for sifra_vzorca in row[10]:
                    row_vzorci.append(vzorci.get(sifra_vzorca, ''))
                    nivo = str(vzorci.get(sifra_vzorca, '')).split("###")
                    try:
                        preisk_vzorci_nivo1.append(nivo[2])
                        preisk_vzorci_nivo2.append(nivo[3])
                        preisk_vzorci_nivo3.append(nivo[4])
                        preisk_vzorci_nivo4.append(nivo[5])
                        preisk_vzorci_nivo5.append(nivo[6])
                        preisk_vzorci_nivo6.append(nivo[7])
                    except IndexError:
                        pass

                if zadetek:
                    zadetek = zadetek[0].getObject()
                    zadetek.podrocje = row[0]
                    zadetek.title = row[2]
                    zadetek.metode = row[3]
                    zadetek.opis = row[4]
                    zadetek.sinonim = row[5]
                    zadetek.trajanje = row[6]
                    zadetek.urnik = row[7]
                    zadetek.opombe = row[8]
                    zadetek.sklop = row[9]
                    zadetek.vzorci_lab_kratica = row[11]
                    zadetek.vzorci_lab = row[12]
                    zadetek.vzorci_lab_tel = row[13]
                    zadetek.vzorci_odvzem = row[14]
                    zadetek.vzorci_odvzem_video_sifra = row[15]
                    zadetek.vzorci_kolicina = row[16]
                    zadetek.embalaza_slika_sifra = row[17]
                    zadetek.vzorci_transport = row[18]
                    zadetek.vzorci_opomba = row[19]
                    zadetek.vzorci = row_vzorci
                    zadetek.vzorci_nivo1 = preisk_vzorci_nivo1
                    zadetek.vzorci_nivo2 = preisk_vzorci_nivo2
                    zadetek.vzorci_nivo3 = preisk_vzorci_nivo3
                    zadetek.vzorci_nivo4 = preisk_vzorci_nivo4
                    zadetek.vzorci_nivo5 = preisk_vzorci_nivo5
                    zadetek.vzorci_nivo6 = preisk_vzorci_nivo6
                    zadetek.laboratoriji = row[20]
                    zadetek.link_sop = row[21]
                    zadetek.nova_pre = row[22]
                    zadetek.nujna_pre = row[23]
                    zadetek.preiskava_vzorec_nujno = row[24]
                    msg('Preiskava ' + novid + ' je bila posodobljena')
                    stevec1 = stevec1 + 1
                else:
                    try:
                        preiskava = self.portal.restrictedTraverse(
                            "portal/preiskave/preiskave-1/katalog-preiskav"
                        ).invokeFactory(
                            type_name='imipreiskava',
                            id=novid,
                            podrocje=row[0],
                            title=row[2],
                            metode=row[3],
                            opis=row[4],
                            sinonim=row[5],
                            trajanje=row[6],
                            urnik=row[7],
                            opombe=row[8],
                            sklop=row[9],
                            vzorci_lab_kratica=row[11],
                            vzorci_lab=row[12],
                            vzorci_lab_tel=row[13],
                            vzorci_odvzem=row[14],
                            vzorci_odvzem_video_sifra=row[15],
                            vzorci_kolicina=row[16],
                            embalaza_slika_sifra=row[17],
                            vzorci_transport=row[18],
                            vzorci_opomba=row[19],
                            vzorci=row_vzorci,
                            vzorci_nivo1=preisk_vzorci_nivo1,
                            vzorci_nivo2=preisk_vzorci_nivo2,
                            vzorci_nivo3=preisk_vzorci_nivo3,
                            vzorci_nivo4=preisk_vzorci_nivo4,
                            vzorci_nivo5=preisk_vzorci_nivo5,
                            vzorci_nivo6=preisk_vzorci_nivo6,
                            laboratoriji=row[20],
                            link_sop=row[21],
                            nova_pre=row[22],
                            nujna_pre=row[23],
                            preiskava_vzorec_nujno=row[24]
                        )
                        stevec2 = stevec2 + 1
                    except BadRequest:
                        msg('Neuspesen vnos: ' + novid)
                        stevec3 = stevec3 + 1

            msg('Posodobljenih vnosov: ' + str(stevec1))
            msg('Ustvarjenih vnosov: ' + str(stevec2))
            msg('Neuspesno obdelanih vrstic: ' + str(stevec3))

            setattr(self, 'datoteka1', None)
            setattr(self, 'datoteka2', None)

    security.declareProtected(ModifyPortalContent, 'at_post_create_script')
    def at_post_create_script(self):
        self.unpackArchive()

    security.declareProtected(ModifyPortalContent, 'at_post_edit_script')
    def at_post_edit_script(self):
        self.unpackArchive()
